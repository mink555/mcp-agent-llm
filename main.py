import os
import json
import time
import argparse
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.loader import BFCLDataLoader
from core.handler import ModelHandler
from core.checker import BFCLChecker
from core.executor import BFCLMockExecutor

def _format_model_name_for_filename(model_name):
    """
    모델명을 파일명에 적합한 약식으로 변환
    
    예시:
    - mistralai/mistral-small-3.2-24b-instruct → mistral_small_3_2_24b
    - openai/gpt-4o-mini → gpt_4o_mini
    - anthropic/claude-3-haiku → claude_3_haiku
    """
    # 회사명/모델명 형식에서 모델명만 추출
    if '/' in model_name:
        model_name = model_name.split('/')[-1]
    
    # 하이픈을 언더바로 변환
    model_name = model_name.replace('-', '_')
    
    # 점을 언더바로 변환
    model_name = model_name.replace('.', '_')
    
    # 불필요한 접미사 제거
    suffixes_to_remove = ['_instruct', '_free', '_turbo', '_preview']
    for suffix in suffixes_to_remove:
        if model_name.endswith(suffix):
            model_name = model_name[:-len(suffix)]
    
    # 소문자 변환
    model_name = model_name.lower()
    
    # 연속된 언더바 제거
    while '__' in model_name:
        model_name = model_name.replace('__', '_')
    
    # 앞뒤 언더바 제거
    model_name = model_name.strip('_')
    
    return model_name

# ==========================================
# [BFCL 공식 카테고리 정의]
# ==========================================
# BFCL V3/V4 전체 카테고리 (공식 벤치마크 구성 - 총 20개)
BFCL_ALL_CATEGORIES = {
    # Single-turn Non-Live (AST Evaluation)
    "simple_python": {"count": 399, "group": "AST_NON_LIVE", "difficulty": "⭐"},
    "simple_javascript": {"count": 49, "group": "AST_NON_LIVE", "difficulty": "⭐"},
    "simple_java": {"count": 99, "group": "AST_NON_LIVE", "difficulty": "⭐"},
    "multiple": {"count": 199, "group": "AST_NON_LIVE", "difficulty": "⭐⭐"},
    "parallel": {"count": 199, "group": "AST_NON_LIVE", "difficulty": "⭐⭐"},
    "parallel_multiple": {"count": 199, "group": "AST_NON_LIVE", "difficulty": "⭐⭐⭐"},
    
    # Single-turn Live (Executable + AST)
    "live_simple": {"count": 257, "group": "AST_LIVE", "difficulty": "⭐⭐"},
    "live_multiple": {"count": 1052, "group": "AST_LIVE", "difficulty": "⭐⭐"},
    "live_parallel": {"count": 15, "group": "AST_LIVE", "difficulty": "⭐⭐⭐"},
    "live_parallel_multiple": {"count": 23, "group": "AST_LIVE", "difficulty": "⭐⭐⭐"},
    
    # Multi-turn (State-based + Response-based)
    "multi_turn_base": {"count": 200, "group": "MULTI_TURN", "difficulty": "⭐⭐⭐"},
    "multi_turn_miss_func": {"count": 200, "group": "MULTI_TURN", "difficulty": "⭐⭐⭐"},
    "multi_turn_miss_param": {"count": 200, "group": "MULTI_TURN", "difficulty": "⭐⭐⭐"},
    "multi_turn_long_context": {"count": 200, "group": "MULTI_TURN", "difficulty": "⭐⭐⭐⭐"},
    
    # Relevance Detection
    "irrelevance": {"count": 239, "group": "RELEVANCE", "difficulty": "⭐⭐"},
    "live_irrelevance": {"count": 884, "group": "RELEVANCE", "difficulty": "⭐⭐"},
    "live_relevance": {"count": 16, "group": "RELEVANCE", "difficulty": "⭐⭐"},
    
    # Agentic (V4 추가)
    "web_search": {"count": 99, "group": "AGENTIC", "difficulty": "⭐⭐⭐"},
    "memory": {"count": 155, "group": "AGENTIC", "difficulty": "⭐⭐⭐"},
    "format_sensitivity": {"count": 9, "group": "AGENTIC", "difficulty": "⭐⭐"},
}

# ==========================================
# [기본 설정]
# ==========================================
DEFAULT_CONFIG = {
    "model_name": "mistralai/mistral-small-3.2-24b-instruct",  # 유료 모델 (tool calling 지원)
    "categories": list(BFCL_ALL_CATEGORIES.keys()),  # 전체 카테고리
    "samples_per_cat": 5,  # 각 카테고리당 기본 샘플 수
    "sampling_strategy": "equal",  # "equal" or "proportional"
    "max_agent_steps": 3,
    "rate_limit_delay": 3  # API 레이트 리밋 대기 시간 (초)
}

# 빠른 테스트용 샘플 설정
QUICK_TEST_CONFIG = {
    "samples_per_cat": 2,
    "categories": ["simple_python", "multiple", "live_simple"],  # 3개 대표 카테고리
    "sampling_strategy": "equal",
    "rate_limit_delay": 5
}

# 전체 벤치마크 설정 (모든 데이터 사용)
FULL_TEST_CONFIG = {
    "samples_per_cat": 999999,  # 각 카테고리의 모든 샘플 사용
    "categories": list(BFCL_ALL_CATEGORIES.keys()),  # 전체 20개 카테고리
    "sampling_strategy": "equal",
    "rate_limit_delay": 3
}

class BFCLScorer:
    """BFCL 표준 점수 산출 클래스"""
    
    @staticmethod
    def calculate_scores(df):
        """BFCL 공식 점수 산출 방법에 따라 통계 계산"""
        scores = {}
        
        # 1. 카테고리별 정확도
        for cat in df['카테고리'].unique():
            cat_df = df[df['카테고리'] == cat]
            pass_count = len(cat_df[cat_df['결과'] == 'PASS'])
            total_count = len(cat_df)
            accuracy = (pass_count / total_count * 100) if total_count > 0 else 0
            scores[cat] = {
                "accuracy": accuracy,
                "pass": pass_count,
                "total": total_count,
                "group": BFCL_ALL_CATEGORIES.get(cat, {}).get("group", "UNKNOWN")
            }
        
        # 2. 그룹별 평균 정확도
        groups = {}
        for cat, data in scores.items():
            group = data["group"]
            if group not in groups:
                groups[group] = []
            groups[group].append(data["accuracy"])
        
        group_scores = {group: sum(accs) / len(accs) if accs else 0 
                       for group, accs in groups.items()}
        
        # 3. Overall 정확도 (unweighted average of all categories)
        all_accuracies = [data["accuracy"] for data in scores.values()]
        overall_accuracy = sum(all_accuracies) / len(all_accuracies) if all_accuracies else 0
        
        return {
            "overall": overall_accuracy,
            "by_category": scores,
            "by_group": group_scores
        }

class ExcelReporter:
    """BFCL 표준 멀티 시트 Excel 리포트 생성기"""
    
    @staticmethod
    def save(df, path, model_name, config):
        """BFCL 표준 4-시트 리포트 생성"""
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            # 1. 상세 결과 시트
            ExcelReporter._write_result_sheet(writer, df)
            
            # 2. 요약 통계 시트 (BFCL 스타일)
            ExcelReporter._write_summary_sheet(writer, df, model_name)
            
            # 3. 데이터셋 정보 시트
            ExcelReporter._write_dataset_info_sheet(writer)
            
            # 4. 참고 자료 시트
            ExcelReporter._write_reference_sheet(writer)
    
    @staticmethod
    def _write_result_sheet(writer, df):
        """상세 결과 시트 (기존 스타일 유지)"""
        # NaN을 빈 문자열로 변환 (Excel 오류 방지)
        df = df.fillna("")
        df.to_excel(writer, index=False, sheet_name='Detailed Results')
        ws = writer.sheets['Detailed Results']
        
        # 스타일 정의
        header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        stripe_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
        pass_font = Font(color='2E7D32', bold=True)
        fail_font = Font(color='D32F2F', bold=True)
        border = Border(
            left=Side(style='thin', color='E0E0E0'), 
            right=Side(style='thin', color='E0E0E0'), 
            top=Side(style='thin', color='E0E0E0'), 
            bottom=Side(style='thin', color='E0E0E0')
        )

        # 헤더 스타일링
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 본문 스타일링
        for row_idx in range(2, len(df) + 2):
            is_stripe = row_idx % 2 == 0
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if is_stripe: cell.fill = stripe_fill
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True, indent=1)
                
                if col_idx == 3: # 결과 컬럼 (PASS/FAIL)
                    cell.font = pass_font if cell.value == "PASS" else fail_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # 너비 설정
        col_widths = [15, 15, 10, 40, 30, 30, 30, 30, 15]
        for i, width in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = width
        ws.freeze_panes = 'C2'
    
    @staticmethod
    def _write_summary_sheet(writer, df, model_name):
        """BFCL 스타일 요약 통계 시트 (한국어 포함, Excel 수식 자동 집계)"""
        scores = BFCLScorer.calculate_scores(df)
        
        # 카테고리별 행 번호 매핑 (Detailed Results 시트 참조용)
        category_rows = {}
        for idx, cat in enumerate(df['카테고리'].unique()):
            # Excel은 1-based, 헤더가 1행이므로 데이터는 2행부터
            cat_df = df[df['카테고리'] == cat]
            first_row = df[df['카테고리'] == cat].index[0] + 2  # +2 for Excel indexing
            last_row = first_row + len(cat_df) - 1
            category_rows[cat] = (first_row, last_row)
        
        # 요약 데이터 구성
        summary_data = []
        
        # 모델 정보
        summary_data.append({
            "지표 (Metric)": "테스트 모델 (Model)",
            "값 (Value)": model_name,
            "설명 (Description)": "평가에 사용된 LLM 모델"
        })
        summary_data.append({
            "지표 (Metric)": " ",
            "값 (Value)": " ",
            "설명 (Description)": " "
        })
        
        # Overall Score (수식으로 계산)
        summary_data.append({
            "지표 (Metric)": "전체 정확도 (Overall Accuracy)",
            "값 (Value)": "FORMULA_OVERALL_ACC",  # 나중에 수식으로 교체
            "설명 (Description)": "모든 카테고리의 비가중 평균 (BFCL 표준)"
        })
        summary_data.append({
            "지표 (Metric)": "📊 점수 산출 공식",
            "값 (Value)": "Σ(Category Acc) / N",
            "설명 (Description)": "N = 카테고리 수, 각 카테고리에 동일한 가중치 부여"
        })
        summary_data.append({
            "지표 (Metric)": " ",
            "값 (Value)": " ",
            "설명 (Description)": " "
        })
        
        # 그룹별 점수 (수식으로 계산)
        summary_data.append({
            "지표 (Metric)": "━━━ 그룹별 점수 (Group Scores) ━━━",
            "값 (Value)": " ",
            "설명 (Description)": " "
        })
        
        group_row_start = len(summary_data) + 2  # 현재까지의 행 수 + 헤더
        
        for group in sorted(set(scores['by_group'].keys())):
            group_kr = {
                "AST_NON_LIVE": "AST 비실행 평가",
                "AST_LIVE": "AST 실행 평가",
                "MULTI_TURN": "멀티턴 대화",
                "RELEVANCE": "관련성 탐지",
                "AGENTIC": "에이전트 기능"
            }.get(group, group)
            summary_data.append({
                "지표 (Metric)": f"{group} ({group_kr})",
                "값 (Value)": f"FORMULA_GROUP_{group}",  # 나중에 수식으로 교체
                "설명 (Description)": f"{group} 그룹 내 카테고리들의 평균 정확도"
            })
        
        summary_data.append({
            "지표 (Metric)": " ",
            "값 (Value)": " ",
            "설명 (Description)": " "
        })
        
        # 카테고리별 상세 점수 (수식으로 계산)
        cat_row_start = len(summary_data) + 2
        
        summary_data.append({
            "지표 (Metric)": "━━━ 카테고리별 점수 (Category Scores) ━━━",
            "값 (Value)": " ",
            "설명 (Description)": " "
        })
        
        for cat, data in scores['by_category'].items():
            cat_kr = ExcelReporter._get_category_name_korean(cat)
            summary_data.append({
                "지표 (Metric)": f"{cat} ({cat_kr})",
                "값 (Value)": f"FORMULA_CAT_{cat}",  # 나중에 수식으로 교체
                "설명 (Description)": f"그룹: {data['group']}, 수식 자동 계산"
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df = summary_df.fillna("")
        summary_df.to_excel(writer, index=False, sheet_name='Summary (BFCL)')
        
        # 스타일링 및 수식 삽입
        ws = writer.sheets['Summary (BFCL)']
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 55
        
        # 스타일 정의
        header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        stripe_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
        section_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )
        
        # 헤더 스타일링
        for col_idx in range(1, 4):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 수식 삽입 및 본문 스타일링
        for row_idx in range(2, len(summary_df) + 2):
            cell_a = ws.cell(row=row_idx, column=1)
            cell_b = ws.cell(row=row_idx, column=2)
            cell_c = ws.cell(row=row_idx, column=3)
            
            # Excel 수식 삽입
            if cell_b.value and isinstance(cell_b.value, str):
                # Overall Accuracy 수식
                if cell_b.value == "FORMULA_OVERALL_ACC":
                    # 카테고리별 점수들의 평균 (퍼센트 값 추출, 소수점 첫째자리까지)
                    cat_start_row = cat_row_start + 1  # 섹션 헤더 다음 행부터
                    cat_end_row = cat_start_row + len(scores['by_category']) - 1
                    # 각 카테고리의 퍼센트 값 추출하여 평균
                    value_extracts = [f'VALUE(LEFT(B{r},FIND("%",B{r})-1))' for r in range(cat_start_row, cat_end_row + 1)]
                    cell_b.value = f'=ROUND(AVERAGE({",".join(value_extracts)}),1)&"%"'
                
                # 그룹별 수식
                elif cell_b.value.startswith("FORMULA_GROUP_"):
                    group = cell_b.value.replace("FORMULA_GROUP_", "")
                    # 해당 그룹에 속하는 카테고리들의 행 번호 찾기
                    group_cats = [cat for cat, data in scores['by_category'].items() if data['group'] == group]
                    if group_cats:
                        cat_rows = []
                        for i, cat in enumerate(scores['by_category'].keys()):
                            if cat in group_cats:
                                cat_rows.append(cat_row_start + 1 + i)
                        
                        if len(cat_rows) == 1:
                            # 퍼센트 값 추출: LEFT(B12, FIND("%", B12)-1)
                            cell_b.value = f'=VALUE(LEFT(B{cat_rows[0]},FIND("%",B{cat_rows[0]})-1))&"%"'
                        else:
                            # 각 카테고리의 퍼센트 값 추출하여 평균 (소수점 첫째자리까지)
                            value_extracts = [f'VALUE(LEFT(B{r},FIND("%",B{r})-1))' for r in cat_rows]
                            cell_b.value = f'=ROUND(AVERAGE({",".join(value_extracts)}),1)&"%"'
                
                # 카테고리별 수식
                elif cell_b.value.startswith("FORMULA_CAT_"):
                    cat = cell_b.value.replace("FORMULA_CAT_", "")
                    if cat in category_rows:
                        first_row, last_row = category_rows[cat]
                        # PASS 개수 / 전체 개수 * 100 (소수점 첫째자리까지)
                        pass_formula = f'COUNTIF(\'Detailed Results\'!C{first_row}:C{last_row},"PASS")'
                        total_formula = f'COUNTA(\'Detailed Results\'!C{first_row}:C{last_row})'
                        cell_b.value = f'=IF({total_formula}=0,0,ROUND({pass_formula}/{total_formula}*100,1))&"%"&" ("&{pass_formula}&"/"&{total_formula}&")"'
            
            # 빈 행은 스타일 없음
            if cell_a.value and str(cell_a.value).strip() == "":
                continue
            # 섹션 구분선 강조
            elif cell_a.value and ("━━━" in str(cell_a.value) or "===" in str(cell_a.value)):
                cell_a.fill = section_fill
                cell_a.font = Font(bold=True, color='333333')
                cell_b.fill = section_fill
                cell_c.fill = section_fill
            # 스트라이프 효과
            elif row_idx % 2 == 0:
                cell_a.fill = stripe_fill
                cell_b.fill = stripe_fill
                cell_c.fill = stripe_fill
            
            # 테두리 및 정렬
            for cell in [cell_a, cell_b, cell_c]:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True, indent=1)
            
            # Value 컬럼 중앙 정렬
            cell_b.alignment = Alignment(horizontal='center', vertical='center')
            
            # Overall Accuracy 강조
            if "전체 정확도" in str(cell_a.value):
                cell_a.font = Font(bold=True, size=11)
                cell_b.font = Font(bold=True, color='2E7D32', size=12)
                cell_b.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.freeze_panes = 'A2'
    
    @staticmethod
    def _write_dataset_info_sheet(writer):
        """데이터셋 정보 시트 (한국어 포함)"""
        dataset_info = []
        
        for cat, info in BFCL_ALL_CATEGORIES.items():
            cat_kr = ExcelReporter._get_category_name_korean(cat)
            dataset_info.append({
                "카테고리 (Category)": f"{cat}\n({cat_kr})",
                "전체 개수\n(Total Count)": info["count"],
                "그룹\n(Group)": info["group"],
                "난이도\n(Difficulty)": info["difficulty"],
                "설명 (Description)": ExcelReporter._get_category_description(cat)
            })
        
        info_df = pd.DataFrame(dataset_info)
        # NaN을 빈 문자열로 변환 (Excel 오류 방지)
        info_df = info_df.fillna("")
        info_df.to_excel(writer, index=False, sheet_name='Dataset Info')
        
        # 스타일링 (Detailed Results 스타일 유지)
        ws = writer.sheets['Dataset Info']
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 60
        
        # 스타일 정의
        header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        stripe_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )
        
        # 헤더 스타일링
        for col_idx in range(1, 6):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # 본문 스타일링
        for row_idx in range(2, len(info_df) + 2):
            is_stripe = row_idx % 2 == 0
            for col_idx in range(1, 6):
                cell = ws.cell(row=row_idx, column=col_idx)
                if is_stripe:
                    cell.fill = stripe_fill
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True, indent=1)
                
                # 중앙 정렬 (Count, Difficulty)
                if col_idx in [2, 4]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.freeze_panes = 'A2'
    
    @staticmethod
    def _write_reference_sheet(writer):
        """참고 자료 시트 (한국어 포함)"""
        references = [
            {"구분 (Section)": "📊 BFCL 공식 (Official)", "내용 (Content)": "Berkeley Function Calling Leaderboard (BFCL)"},
            {"구분 (Section)": "🌐 웹사이트 (Website)", "내용 (Content)": "https://gorilla.cs.berkeley.edu/leaderboard.html"},
            {"구분 (Section)": "💻 GitHub", "내용 (Content)": "https://github.com/ShishirPatil/gorilla/tree/main/berkeley-function-call-leaderboard"},
            {"구분 (Section)": "📁 Dataset", "내용 (Content)": "https://huggingface.co/datasets/gorilla-llm/Berkeley-Function-Calling-Leaderboard"},
            {"구분 (Section)": " ", "내용 (Content)": " "},
            {"구분 (Section)": "━━━ 평가 방법 (Evaluation Methods) ━━━", "내용 (Content)": " "},
            {"구분 (Section)": "🌳 AST 평가", "내용 (Content)": "추상 구문 트리 비교 (Abstract Syntax Tree comparison for structural correctness)"},
            {"구분 (Section)": "⚙️ 실행 평가", "내용 (Content)": "REST API 및 Python 함수 실제 실행 (Actual execution for REST APIs and Python functions)"},
            {"구분 (Section)": "✅ 관련성 탐지", "내용 (Content)": "관련 없는 함수 호출 회피 능력 (Ability to avoid irrelevant function calls)"},
            {"구분 (Section)": "🔀 병렬 호출 순서", "내용 (Content)": "BFCL 표준: 병렬(parallel) 카테고리는 호출 순서 무시. 집합처럼 매칭 (Order-independent matching for parallel function calls)"},
            {"구분 (Section)": " ", "내용 (Content)": " "},
            {"구분 (Section)": "━━━ 점수 산출 (Scoring) ━━━", "내용 (Content)": " "},
            {"구분 (Section)": "📈 전체 정확도", "내용 (Content)": "Overall Accuracy = Σ(Category Accuracy) / N (모든 카테고리의 비가중 평균)"},
            {"구분 (Section)": "📊 카테고리 정확도", "내용 (Content)": "Category Accuracy = (PASS count / Total count) × 100%"},
            {"구분 (Section)": "📂 그룹 정확도", "내용 (Content)": "Group Accuracy = 동일 그룹 내 카테고리들의 평균 (Average of categories within same group)"},
            {"구분 (Section)": " ", "내용 (Content)": " "},
            {"구분 (Section)": "━━━ 논문 (Papers) ━━━", "내용 (Content)": " "},
            {"구분 (Section)": "📄 BFCL v1", "내용 (Content)": "AST 평가 메트릭 도입 (Introducing AST evaluation metric)"},
            {"구분 (Section)": "📄 BFCL v2", "내용 (Content)": "기업 및 OSS 기여 함수 (Enterprise and OSS-contributed functions)"},
            {"구분 (Section)": "📄 BFCL v3", "내용 (Content)": "멀티턴 상호작용 (Multi-turn interactions)"},
            {"구분 (Section)": "📄 BFCL v4", "내용 (Content)": "종합적 에이전트 평가 (Holistic agentic evaluation)"},
        ]
        
        ref_df = pd.DataFrame(references)
        # NaN을 빈 문자열로 변환 (Excel 오류 방지)
        ref_df = ref_df.fillna("")
        ref_df.to_excel(writer, index=False, sheet_name='Reference')
        
        # 스타일링 (Detailed Results 스타일 유지)
        ws = writer.sheets['Reference']
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 85
        
        # 스타일 정의
        header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        stripe_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
        section_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )
        
        # 헤더 스타일링
        for col_idx in range(1, 3):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 본문 스타일링
        for row_idx in range(2, len(ref_df) + 2):
            cell_a = ws.cell(row=row_idx, column=1)
            cell_b = ws.cell(row=row_idx, column=2)
            
            # 빈 행은 스타일 없음
            if cell_a.value and str(cell_a.value).strip() == "":
                continue
            # 섹션 구분선 강조
            elif cell_a.value and "━━━" in str(cell_a.value):
                cell_a.fill = section_fill
                cell_a.font = Font(bold=True, color='333333')
                cell_b.fill = section_fill
            # 스트라이프 효과
            elif row_idx % 2 == 0:
                cell_a.fill = stripe_fill
                cell_b.fill = stripe_fill
            
            # 테두리 및 정렬
            for cell in [cell_a, cell_b]:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True, indent=1)
        
        ws.freeze_panes = 'A2'
    
    @staticmethod
    def _get_category_name_korean(cat):
        """카테고리 한국어 이름 반환"""
        korean_names = {
            "simple_python": "단일 Python",
            "simple_javascript": "단일 JavaScript",
            "simple_java": "단일 Java",
            "multiple": "다중 파라미터",
            "parallel": "병렬 호출",
            "parallel_multiple": "병렬 다중",
            "live_simple": "실행 단일",
            "live_multiple": "실행 다중",
            "live_parallel": "실행 병렬",
            "live_parallel_multiple": "실행 병렬 다중",
            "multi_turn_base": "멀티턴 기본",
            "multi_turn_miss_func": "멀티턴 함수누락",
            "multi_turn_miss_param": "멀티턴 파라미터누락",
            "multi_turn_long_context": "멀티턴 긴컨텍스트",
            "irrelevance": "관련없음 탐지",
            "live_irrelevance": "실행 관련없음",
            "live_relevance": "실행 관련성",
            "web_search": "웹 검색",
            "memory": "메모리 관리",
            "format_sensitivity": "포맷 민감도",
        }
        return korean_names.get(cat, cat)
    
    @staticmethod
    def _get_category_description(cat):
        """카테고리별 설명 반환 (한국어 + 영어)"""
        descriptions = {
            "simple_python": "기본 파라미터를 가진 단일 Python 함수 호출 | Single Python function call with basic parameters",
            "simple_javascript": "기본 파라미터를 가진 단일 JavaScript 함수 호출 | Single JavaScript function call with basic parameters",
            "simple_java": "기본 파라미터를 가진 단일 Java 함수 호출 | Single Java function call with basic parameters",
            "multiple": "단일 함수에 여러 파라미터 전달 | Multiple parameters in single function call",
            "parallel": "여러 함수를 병렬로 호출 (순서 무시) | Multiple functions called in parallel (order-independent)",
            "parallel_multiple": "여러 함수를 다중 파라미터로 병렬 호출 (순서 무시) | Multiple functions with multiple parameters in parallel (order-independent)",
            "live_simple": "실제 API를 사용한 단순 함수 호출 | Simple function calls using live/real APIs",
            "live_multiple": "실제 API에 다중 파라미터 함수 | Multiple parameter functions with live APIs",
            "live_parallel": "실제 API 병렬 함수 호출 (순서 무시) | Parallel function calls with live APIs (order-independent)",
            "live_parallel_multiple": "실제 API 복잡한 병렬 호출 (순서 무시) | Complex parallel calls with live APIs (order-independent)",
            "multi_turn_base": "기본 멀티턴 대화형 함수 호출 | Basic multi-turn conversational function calling",
            "multi_turn_miss_func": "함수 누락 처리가 필요한 멀티턴 | Multi-turn with missing function handling",
            "multi_turn_miss_param": "파라미터 누락 처리가 필요한 멀티턴 | Multi-turn with missing parameter handling",
            "multi_turn_long_context": "긴 컨텍스트가 필요한 멀티턴 | Multi-turn with extended context requirements",
            "irrelevance": "함수를 호출하지 말아야 할 때 탐지 | Detecting when NOT to call any function",
            "live_irrelevance": "실제 API에서 관련없는 함수 호출 회피 | Avoiding irrelevant function calls with live APIs",
            "live_relevance": "실제 API에서 관련 있는 함수 탐지 | Detecting relevant functions with live APIs",
            "web_search": "웹 검색 에이전트 기능 | Web search agent capabilities",
            "memory": "대화 메모리 관리 및 컨텍스트 유지 | Memory management and context retention",
            "format_sensitivity": "다양한 입력 포맷에 대한 민감도 | Sensitivity to various input formats",
        }
        return descriptions.get(cat, "설명 없음 | No description available")

def process_test_case(handler, executor, checker, cat, q, a, max_steps=3):
    """단일 테스트 케이스 처리 로직"""
    test_id = q['id']
    tools = BFCLDataLoader().get_functions(cat, q)
    gt = a['ground_truth']
    
    # 개선된 시스템 프롬프트 (Tool Calling Best Practices 적용)
    SYSTEM_PROMPT = """You are an expert function-calling assistant. Your primary job is to call the appropriate functions with correct parameters.

CRITICAL RULES:
1. ALWAYS use the provided functions to answer user requests - this is your main purpose
2. NEVER make up or hallucinate function names - only use functions from the tools list
3. Extract parameter values directly from the user's question
4. For multi-step tasks, call functions sequentially and use their results

PARAMETER EXTRACTION:
- Read the user's question carefully to extract all required parameter values
- Use exact numbers, strings, or values provided by the user
- Follow parameter type specifications (string, number, boolean, array, object)
- If a parameter format is specified (e.g., "City, State"), follow it exactly

FUNCTION SELECTION:
- Match the user's intent to the most appropriate function name
- Check function descriptions to understand their purpose
- Consider function parameters to ensure you have the required data

MULTI-TURN BEHAVIOR:
- Use tool execution results to inform your next function call
- Chain multiple function calls when needed to complete complex tasks
- Interpret tool responses and extract relevant information for subsequent calls

Your goal is to successfully call the right functions with the right parameters."""

    # 멀티턴 질문 구조화 (유연한 대응)
    raw_question = q['question']
    if isinstance(raw_question[0], list):
        user_turns = raw_question
    else:
        user_turns = [[{"role": "user", "content": msg} for msg in raw_question]]

    messages = []
    all_model_calls = []
    final_res = None
    final_content = ""

    for turn_idx, turn_msgs in enumerate(user_turns):
        messages.extend(turn_msgs)
        
        # 에이전트 루프 (멀티홉 처리)
        for step in range(max_steps):
            res = handler.inference(
                messages=[{"role": "system", "content": SYSTEM_PROMPT}] + messages,
                tools=tools,
                temperature=0,
                force_tool=(cat not in ["irrelevance", "multi_turn_miss_func"])
            )
            
            final_res = res
            final_content = res["content"]
            ast_out = handler.decode_ast(res)

            if ast_out:
                all_model_calls.extend(ast_out)
                messages.append(res["msg_obj"]) # 어시스턴트 메시지 추가
                
                # 도구 실행 및 결과 추가
                for i, call in enumerate(ast_out):
                    feedback = executor.execute(call)
                    # Tool ID 매칭 (OpenAI 호환)
                    call_id = res["msg_obj"].tool_calls[i].id if (res["msg_obj"].tool_calls and len(res["msg_obj"].tool_calls) > i) else f"call_{turn_idx}_{step}_{i}"
                    messages.append({"role": "tool", "tool_call_id": call_id, "content": feedback})
                
                # 카테고리별 루프 전략
                # Simple/Single-turn 카테고리: 첫 번째 도구 호출 후 종료
                if cat in ["simple_python", "simple_javascript", "simple_java", "live_simple", "web_search"]:
                    break
                # Multiple/Parallel: 여러 도구를 한 번에 호출 후 종료 (Live 포함)
                elif cat in ["multiple", "parallel", "parallel_multiple", 
                             "live_multiple", "live_parallel", "live_parallel_multiple"]:
                    break
                # Multi-turn 카테고리: 다음 사용자 턴으로 이동
                elif "multi_turn" in cat:
                    break
                # Relevance 카테고리: 단일 판단이므로 break
                elif cat in ["irrelevance", "live_irrelevance", "live_relevance"]:
                    break
                # Agentic 카테고리 (memory, format_sensitivity): 멀티홉 가능성 있음
                else:
                    continue
            else:
                messages.append({"role": "assistant", "content": res["content"]})
                break # 도구 호출 없으면 루프 종료

    # 최종 검증
    is_pass, detail = checker.ast_checker(all_model_calls, gt, final_content, cat)
    
    return {
        "카테고리": cat, "ID": test_id, "결과": "PASS" if is_pass else "FAIL",
        "질문": str(user_turns[0][0]['content']), "검증 상세": detail,
        "사고과정": final_res["thinking"] if final_res else "N/A",
        "누적 호출(AST)": json.dumps(all_model_calls, ensure_ascii=False),
        "정답(GT)": json.dumps(gt, ensure_ascii=False),
        "Latency": final_res["latency"] if final_res else 0
    }

def run_benchmark(config):
    """
    벤치마크 실행 함수
    
    Args:
        config (dict): 벤치마크 설정 딕셔너리
            - model_name: 모델 이름
            - categories: 테스트할 카테고리 리스트
            - samples_per_cat: 카테고리당 샘플 수
            - max_agent_steps: 최대 에이전트 스텝
            - rate_limit_delay: API 호출 간 대기 시간
    """
    # API 키 확인
    load_dotenv()
    api_key = os.getenv("OPENROUTER_API_KEY")
    if not api_key:
        raise ValueError("OPENROUTER_API_KEY가 설정되지 않았습니다. .env 파일을 확인해주세요.")
    
    loader = BFCLDataLoader()
    handler = ModelHandler(api_key=api_key, model_name=config["model_name"])
    checker = BFCLChecker()
    
    all_results = []
    total_samples = len(config["categories"]) * config["samples_per_cat"]
    
    print("=" * 80)
    print(f"🚀 BFCL 벤치마크 시작")
    print("=" * 80)
    print(f"📋 모델: {config['model_name']}")
    print(f"📂 카테고리: {', '.join(config['categories'])}")
    print(f"📊 카테고리당 샘플: {config['samples_per_cat']}개")
    print(f"🎯 총 예상 테스트: {total_samples}개")
    print("=" * 80)

    start_time = time.time()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    mode_tag = "QUICK" if config["samples_per_cat"] <= 2 else "FULL"
    model_short = _format_model_name_for_filename(config["model_name"])
    os.makedirs("results", exist_ok=True)
    
    for cat_idx, cat in enumerate(config["categories"], 1):
        cat_results = []
        print(f"\n[{cat_idx}/{len(config['categories'])}] 📂 Category: {cat}")
        questions, answers = loader.load_dataset(cat, limit=config["samples_per_cat"])
        
        if not questions:
            print(f"  ⚠️  데이터 없음, 스킵")
            continue

        for idx, (q, a) in enumerate(zip(questions, answers), 1):
            print(f"  [{idx}/{len(questions)}] Testing: {q['id'][:30]}...", end=" ")
            executor = BFCLMockExecutor(initial_config=q.get('initial_config'))
            
            try:
                result = process_test_case(
                    handler, executor, checker, cat, q, a, 
                    max_steps=config["max_agent_steps"]
                )
                all_results.append(result)
                cat_results.append(result)
                status = "✅" if result["결과"] == "PASS" else "❌"
                print(f"{status} ({result['Latency']:.0f}ms)")
            except Exception as e:
                print(f"❌ ERROR: {str(e)[:50]}")
                continue
        
        # 카테고리별 결과 저장
        if cat_results:
            cat_df = pd.DataFrame(cat_results)
            report_path = f"results/BFCL_{mode_tag}_{model_short}_{cat}_Report_{timestamp}.xlsx"
            ExcelReporter.save(cat_df, report_path, config["model_name"], config)
            
            cat_pass = len(cat_df[cat_df['결과'] == 'PASS'])
            cat_total = len(cat_df)
            cat_acc = (cat_pass / cat_total * 100) if cat_total > 0 else 0
            print(f"  💾 저장됨: {report_path} ({cat_pass}/{cat_total}, {cat_acc:.1f}%)")
        
        # 레이트 리밋 방지 대기 (마지막 카테고리는 제외)
        if cat_idx < len(config["categories"]):
            print(f"  ⏳ {config['rate_limit_delay']}초 대기 중...")
            time.sleep(config["rate_limit_delay"])

    # 전체 결과 통계
    if not all_results:
        print("\n❌ 결과가 없습니다. 벤치마크를 확인해주세요.")
        return None
    
    elapsed = time.time() - start_time
    df = pd.DataFrame(all_results)
    pass_count = len(df[df['결과'] == 'PASS'])
    total_count = len(df)
    accuracy = (pass_count / total_count * 100) if total_count > 0 else 0
    
    print("\n" + "=" * 80)
    print("✅ 벤치마크 완료!")
    print("=" * 80)
    print(f"📊 총 테스트: {total_count}개")
    print(f"✅ PASS: {pass_count}개 ({accuracy:.1f}%)")
    print(f"❌ FAIL: {total_count - pass_count}개")
    print(f"⏱️  소요 시간: {elapsed:.1f}초")
    print(f"📁 저장 폴더: results/ ({len(config['categories'])}개 파일)")
    print("=" * 80)
    
    return f"results/BFCL_{mode_tag}_{model_short}_*_Report_{timestamp}.xlsx"

def main():
    """메인 실행 함수"""
    parser = argparse.ArgumentParser(
        description="BFCL Benchmark Runner - Function Calling 벤치마크 실행",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
사용 예시:
  # 빠른 샘플 테스트 (3개 카테고리 x 2개 샘플 = 6개)
  python main.py --quick
  
  # 전체 실행 (모든 카테고리 x 모든 샘플 = ~4,693개)
  python main.py --full
  
  # 커스텀 설정
  python main.py --samples 3 --categories simple_python multiple
  
  # 특정 모델로 실행
  python main.py --model "anthropic/claude-3-haiku" --samples 2
  
  # 대기 시간을 줄여서 빠르게 실행
  python main.py --quick --delay 1
        """
    )
    
    parser.add_argument(
        "--quick", 
        action="store_true",
        help="빠른 테스트 모드 (2개 카테고리, 각 2개 샘플)"
    )
    
    parser.add_argument(
        "--full",
        action="store_true",
        help="전체 벤치마크 실행 (모든 카테고리, 모든 샘플 ~4,693개)"
    )
    
    parser.add_argument(
        "--samples",
        type=int,
        help="카테고리당 샘플 수 (기본값: 5)"
    )
    
    parser.add_argument(
        "--categories",
        nargs="+",
        help="테스트할 카테고리 리스트 (예: simple_python multiple)"
    )
    
    parser.add_argument(
        "--model",
        type=str,
        help="사용할 모델 이름 (기본값: mistral-small-3.1)"
    )
    
    parser.add_argument(
        "--delay",
        type=int,
        help="카테고리 간 대기 시간 (초, 기본값: quick=5, full=3, default=3)"
    )
    
    args = parser.parse_args()
    
    # 설정 구성
    if args.quick:
        config = {**DEFAULT_CONFIG, **QUICK_TEST_CONFIG}
        print("🚀 빠른 테스트 모드 실행 (3개 카테고리 × 2개 샘플 = 6개)\n")
    elif args.full:
        config = {**DEFAULT_CONFIG, **FULL_TEST_CONFIG}
        print("🚀 전체 벤치마크 모드 실행 (20개 카테고리 × 모든 샘플 = ~4,693개)\n")
    else:
        config = DEFAULT_CONFIG.copy()
        
    # 커맨드라인 인자로 오버라이드
    if args.samples:
        config["samples_per_cat"] = args.samples
    if args.categories:
        config["categories"] = args.categories
    if args.model:
        config["model_name"] = args.model
    if args.delay is not None:
        config["rate_limit_delay"] = args.delay
    
    # 벤치마크 실행
    run_benchmark(config)

if __name__ == "__main__":
    main()
